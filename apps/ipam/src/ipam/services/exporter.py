from __future__ import annotations

import csv
import io

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from ipam.models import IPAssignment, Network, Site, ValidationIssue


def _join(values: list[str] | None) -> str:
    return ", ".join(values or [])


def export_assignments_csv(session: Session) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "site",
        "network",
        "cidr",
        "vlan",
        "ip_address",
        "hostname",
        "interface",
        "role",
        "status",
        "gateway",
        "dns_servers",
        "search_domains",
        "notes",
    ])
    assignments = session.scalars(
        select(IPAssignment).options(selectinload(IPAssignment.site), selectinload(IPAssignment.network)).order_by(IPAssignment.ip_address)
    ).all()
    for assignment in assignments:
        writer.writerow([
            assignment.site.name if assignment.site else "",
            assignment.network.name if assignment.network else "",
            assignment.network.cidr if assignment.network else "",
            assignment.network.vlan_id if assignment.network else "",
            assignment.ip_address,
            assignment.hostname or "",
            assignment.interface_name or "",
            assignment.role or "",
            assignment.status,
            assignment.gateway or "",
            _join(assignment.dns_servers),
            _join(assignment.search_domains),
            assignment.description or "",
        ])
    return output.getvalue().encode("utf-8")


def export_workbook(session: Session) -> bytes:
    workbook = Workbook()

    sites_ws = workbook.active
    sites_ws.title = "Sites"
    sites_ws.append(["name", "code", "location", "environment", "notes"])
    for site in session.scalars(select(Site).order_by(Site.name)):
        sites_ws.append([site.name, site.code or "", site.location or "", site.environment or "", site.notes or ""])

    networks_ws = workbook.create_sheet("Networks")
    networks_ws.append(["site", "name", "cidr", "vlan_id", "gateway", "dns_servers", "search_domains", "notes"])
    for network in session.scalars(select(Network).options(selectinload(Network.site)).order_by(Network.cidr, Network.name)):
        networks_ws.append([
            network.site.name if network.site else "",
            network.name or "",
            network.cidr or "",
            network.vlan_id or "",
            network.gateway or "",
            _join(network.dns_servers),
            _join(network.search_domains),
            network.notes or "",
        ])

    assignments_ws = workbook.create_sheet("Assignments")
    assignments_ws.append([
        "site",
        "network",
        "cidr",
        "ip_address",
        "hostname",
        "interface_name",
        "role",
        "status",
        "gateway",
        "dns_servers",
        "search_domains",
        "description",
    ])
    for assignment in session.scalars(
        select(IPAssignment).options(selectinload(IPAssignment.site), selectinload(IPAssignment.network)).order_by(IPAssignment.ip_address)
    ):
        assignments_ws.append([
            assignment.site.name if assignment.site else "",
            assignment.network.name if assignment.network else "",
            assignment.network.cidr if assignment.network else "",
            assignment.ip_address,
            assignment.hostname or "",
            assignment.interface_name or "",
            assignment.role or "",
            assignment.status,
            assignment.gateway or "",
            _join(assignment.dns_servers),
            _join(assignment.search_domains),
            assignment.description or "",
        ])

    issues_ws = workbook.create_sheet("Issues")
    issues_ws.append(["category", "severity", "title", "details", "resource_type", "resource_id", "created_at"])
    for issue in session.scalars(select(ValidationIssue).order_by(ValidationIssue.created_at.desc())):
        issues_ws.append([
            issue.category,
            issue.severity,
            issue.title,
            issue.details or "",
            issue.resource_type or "",
            issue.resource_id or "",
            issue.created_at.isoformat(),
        ])

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
