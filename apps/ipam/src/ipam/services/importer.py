from __future__ import annotations

from dataclasses import dataclass
import ipaddress
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import delete, func, select
from sqlalchemy.orm import Session

from ipam.models import IPAssignment, ImportBatch, Network, Site


CANONICAL_HEADERS = {
    "site_name": ["site", "site name", "location", "facility", "datacenter", "office", "campus"],
    "site_code": ["site code", "code"],
    "environment": ["environment", "env"],
    "network_name": ["network name", "segment", "segment name", "vlan name", "network label"],
    "network_address": ["network", "subnet", "network address", "subnet address", "range"],
    "cidr": ["cidr", "subnet/cidr", "network cidr", "subnet cidr", "subnet id"],
    "netmask": ["netmask", "subnet mask", "mask", "prefix"],
    "vlan_id": ["vlan", "vlan id", "vlan number"],
    "ip_address": [
        "ip",
        "ip address",
        "address",
        "host ip",
        "device ip",
        "node ip",
        "nodeip",
        "bind address",
        "advertise address",
    ],
    "hostname": ["hostname", "host name", "device name", "node name", "server name", "fqdn"],
    "interface_name": ["interface", "nic", "adapter", "port", "eth", "interface name"],
    "role": ["role", "type", "node role", "device role", "function"],
    "status": ["status", "state", "assignment status", "usage"],
    "gateway": ["gateway", "default gateway"],
    "dns_servers": ["dns", "dns servers", "name servers", "dns server"],
    "search_domains": ["search domains", "search domain", "domains"],
    "label": ["description", "label", "name"],
    "host_range": ["host address range", "address range"],
    "broadcast_address": ["broadcast address"],
    "notes": ["notes", "comments", "details", "remarks"],
}

RECOGNIZED_HEADERS = {
    normalized
    for aliases in CANONICAL_HEADERS.values()
    for normalized in aliases
}

NON_EMPTY_CARRY_FIELDS = {
    "site_name",
    "site_code",
    "environment",
    "network_name",
    "network_address",
    "cidr",
    "netmask",
    "vlan_id",
    "gateway",
    "dns_servers",
    "search_domains",
}


@dataclass
class SheetImportSummary:
    sheet_name: str
    header_row: int | None
    headers: list[str]
    total_rows: int
    imported_rows: int
    skipped_rows: int


@dataclass
class ImportSummary:
    batch: ImportBatch
    sheet_summaries: list[SheetImportSummary]

    @property
    def imported_rows(self) -> int:
        return sum(item.imported_rows for item in self.sheet_summaries)

    @property
    def total_rows(self) -> int:
        return sum(item.total_rows for item in self.sheet_summaries)


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[_\-/]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text


def clean_cell(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


def split_multi_value(value: str | None) -> list[str]:
    if not value:
        return []
    stripped = value.strip().strip("[]")
    parts = re.split(r"[;,\n]+", stripped)
    return [part.strip().strip('"').strip("'") for part in parts if part.strip()]


def canonical_field(header_value: str) -> str | None:
    normalized = normalize_header(header_value)
    for field_name, aliases in CANONICAL_HEADERS.items():
        if normalized in aliases:
            return field_name
    return None


def looks_like_address(value: str) -> bool:
    try:
        ipaddress.ip_address(value)
        return True
    except ValueError:
        return False


def normalize_ip(value: str | None) -> str | None:
    if not value:
        return None
    try:
        return str(ipaddress.ip_address(value.strip()))
    except ValueError:
        return None


def looks_like_netmask(value: str | None) -> bool:
    if not value:
        return False
    try:
        parsed = ipaddress.ip_address(value.strip())
    except ValueError:
        return False
    return str(parsed).startswith("255.") or str(parsed) == "0.0.0.0"


def normalize_cidr(network_value: str | None, cidr_value: str | None, netmask_value: str | None) -> str | None:
    candidate = (cidr_value or network_value or "").strip()
    if candidate and "/" in candidate:
        try:
            return str(ipaddress.ip_network(candidate, strict=False))
        except ValueError:
            pass

    if network_value and netmask_value:
        network_text = network_value.strip()
        mask_text = netmask_value.strip()
        try:
            if mask_text.isdigit():
                return str(ipaddress.ip_network(f"{network_text}/{mask_text}", strict=False))
            return str(ipaddress.ip_network(f"{network_text}/{mask_text}", strict=False))
        except ValueError:
            return None

    if network_value:
        try:
            parsed = ipaddress.ip_network(network_value.strip(), strict=False)
            return str(parsed)
        except ValueError:
            return None
    return None


def find_header_row(worksheet) -> tuple[int | None, dict[int, str]]:
    best_row = None
    best_mapping: dict[int, str] = {}
    best_score = -1
    max_scan = min(15, worksheet.max_row)

    for row_index, row in enumerate(worksheet.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        mapping: dict[int, str] = {}
        score = 0
        for cell_index, cell_value in enumerate(row):
            field_name = canonical_field(clean_cell(cell_value) or "")
            if field_name:
                mapping[cell_index] = field_name
                score += 1
        if score > best_score:
            best_row = row_index
            best_mapping = mapping
            best_score = score

    if best_score <= 0:
        return None, {}
    return best_row, best_mapping


def meaningful_row(row_data: dict[str, Any]) -> bool:
    return any(
        row_data.get(field)
        for field in (
            "ip_address",
            "hostname",
            "network_address",
            "cidr",
            "network_name",
            "site_name",
            "vlan_id",
            "host_range",
            "label",
        )
    )


def apply_fill_down(row_data: dict[str, Any], previous_values: dict[str, Any]) -> dict[str, Any]:
    merged = dict(row_data)
    for field_name in NON_EMPTY_CARRY_FIELDS:
        if not merged.get(field_name) and previous_values.get(field_name):
            merged[field_name] = previous_values[field_name]
        elif merged.get(field_name):
            previous_values[field_name] = merged[field_name]
    return merged


def _site_lookup(session: Session, site_name: str) -> Site | None:
    return session.scalar(select(Site).where(func.lower(Site.name) == site_name.lower()))


GENERIC_SITE_TOKENS = {
    "external",
    "internal",
    "reserved",
    "spare",
    "cluster",
    "roof",
    "gateway",
    "p2svpn",
    "gatewaysubnet",
    "gatewayaddress",
    "azure",
    "azurebastionsubnet",
}


def infer_site_name(*candidates: str | None, default: str) -> str:
    for candidate in candidates:
        if not candidate:
            continue
        token = re.split(r"[_\-\s]", candidate.strip().lower(), maxsplit=1)[0]
        if token and token not in GENERIC_SITE_TOKENS and len(token) >= 3:
            return token.upper()
    return default


def get_or_create_site(session: Session, row_data: dict[str, Any], default_site_name: str) -> Site:
    site_name = row_data.get("site_name") or default_site_name
    site = _site_lookup(session, site_name)
    if site is None:
        site = Site(
            name=site_name,
            code=row_data.get("site_code"),
            location=row_data.get("site_name") or default_site_name,
            environment=row_data.get("environment"),
            notes=row_data.get("notes"),
        )
        session.add(site)
        session.flush()
        return site

    if row_data.get("site_code") and not site.code:
        site.code = row_data["site_code"]
    if row_data.get("environment") and not site.environment:
        site.environment = row_data["environment"]
    if row_data.get("notes") and not site.notes:
        site.notes = row_data["notes"]
    if row_data.get("site_name") and not site.location:
        site.location = row_data["site_name"]
    return site


def get_or_create_network(
    session: Session,
    *,
    site: Site,
    batch: ImportBatch,
    source_file: str,
    source_sheet: str,
    source_row: int,
    row_data: dict[str, Any],
    network_cache: dict[tuple, Network],
) -> Network | None:
    cidr = normalize_cidr(row_data.get("network_address"), row_data.get("cidr"), row_data.get("netmask"))
    network_name = row_data.get("network_name")
    vlan_id = row_data.get("vlan_id")

    if not any([cidr, network_name, vlan_id]):
        return None

    cache_key = (site.id, cidr or "", (network_name or "").lower(), vlan_id or "")
    if cache_key in network_cache:
        return network_cache[cache_key]

    network = session.scalar(
        select(Network).where(
            Network.site_id == site.id,
            func.coalesce(Network.cidr, "") == (cidr or ""),
            func.coalesce(func.lower(Network.name), "") == ((network_name or "").lower()),
            func.coalesce(Network.vlan_id, "") == (vlan_id or ""),
        )
    )
    if network is None:
        network = Network(
            site_id=site.id,
            source_batch_id=batch.id,
            source_file=source_file,
            source_sheet=source_sheet,
            source_row=source_row,
            name=network_name,
            cidr=cidr,
            vlan_id=vlan_id,
            gateway=normalize_ip(row_data.get("gateway")) or row_data.get("gateway"),
            dns_servers=split_multi_value(row_data.get("dns_servers")),
            search_domains=split_multi_value(row_data.get("search_domains")),
            notes=row_data.get("notes"),
            source_payload=row_data,
        )
        session.add(network)
        session.flush()
    else:
        if not network.gateway and row_data.get("gateway"):
            network.gateway = normalize_ip(row_data.get("gateway")) or row_data.get("gateway")
        if not network.dns_servers and row_data.get("dns_servers"):
            network.dns_servers = split_multi_value(row_data.get("dns_servers"))
        if not network.search_domains and row_data.get("search_domains"):
            network.search_domains = split_multi_value(row_data.get("search_domains"))
        if not network.notes and row_data.get("notes"):
            network.notes = row_data.get("notes")

    network_cache[cache_key] = network
    return network


def import_workbook(session: Session, workbook_path: Path) -> ImportSummary:
    workbook = load_workbook(workbook_path, data_only=True)
    filename = workbook_path.name

    session.execute(delete(IPAssignment).where(IPAssignment.source_file == filename))
    session.execute(delete(Network).where(Network.source_file == filename))
    session.flush()

    batch = ImportBatch(filename=filename, original_path=str(workbook_path), sheet_names=workbook.sheetnames)
    session.add(batch)
    session.flush()

    summaries: list[SheetImportSummary] = []
    network_cache: dict[tuple, Network] = {}

    for worksheet in workbook.worksheets:
        header_row, header_mapping = find_header_row(worksheet)
        if header_row is None:
            summaries.append(
                SheetImportSummary(
                    sheet_name=worksheet.title,
                    header_row=None,
                    headers=[],
                    total_rows=max(worksheet.max_row - 1, 0),
                    imported_rows=0,
                    skipped_rows=max(worksheet.max_row - 1, 0),
                )
            )
            continue

        header_values = [clean_cell(value) or "" for value in next(worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
        context_values = []
        if header_row > 1:
            context_values = [clean_cell(value) for value in next(worksheet.iter_rows(min_row=header_row - 1, max_row=header_row - 1, values_only=True))]
        sheet_context = None
        if context_values and context_values[0] and all(not value for value in context_values[1:]) and not looks_like_netmask(context_values[0]):
            sheet_context = context_values[0]
        previous_values: dict[str, Any] = {}
        if sheet_context:
            previous_values["environment"] = sheet_context
        imported_rows = 0
        skipped_rows = 0
        current_network: Network | None = None

        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            row_data: dict[str, Any] = {}
            for cell_index, field_name in header_mapping.items():
                if cell_index < len(row):
                    row_data[field_name] = clean_cell(row[cell_index])

            row_data = apply_fill_down(row_data, previous_values)
            if sheet_context and not row_data.get("environment"):
                row_data["environment"] = sheet_context
            if not meaningful_row(row_data):
                skipped_rows += 1
                continue

            if not row_data.get("network_name") and row_data.get("cidr") and row_data.get("label"):
                row_data["network_name"] = row_data.get("label")
            if not row_data.get("hostname") and not row_data.get("cidr") and row_data.get("label"):
                row_data["hostname"] = row_data.get("label")
            if not row_data.get("ip_address") and row_data.get("host_range"):
                candidate_ip = normalize_ip(row_data.get("host_range"))
                if candidate_ip:
                    row_data["ip_address"] = candidate_ip

            network_cidr = normalize_cidr(row_data.get("network_address"), row_data.get("cidr"), row_data.get("netmask"))
            is_network_row = bool(network_cidr)
            default_site_name = worksheet.title
            if current_network and current_network.site:
                default_site_name = current_network.site.name
            if sheet_context and not looks_like_netmask(sheet_context):
                default_site_name = sheet_context

            derived_site_name = infer_site_name(
                row_data.get("site_name"),
                row_data.get("hostname"),
                row_data.get("network_name"),
                row_data.get("label"),
                default=default_site_name,
            )
            row_data["site_name"] = derived_site_name

            site = get_or_create_site(session, row_data, derived_site_name)
            network = None
            if is_network_row:
                network = get_or_create_network(
                    session,
                    site=site,
                    batch=batch,
                    source_file=filename,
                    source_sheet=worksheet.title,
                    source_row=row_index,
                    row_data=row_data,
                    network_cache=network_cache,
                )
                current_network = network
            elif current_network is not None:
                network = current_network

            ip_address = normalize_ip(row_data.get("ip_address"))
            if not ip_address:
                if network is not None:
                    imported_rows += 1
                else:
                    skipped_rows += 1
                continue

            assignment = IPAssignment(
                site_id=site.id,
                network_id=network.id if network else None,
                source_batch_id=batch.id,
                source_file=filename,
                source_sheet=worksheet.title,
                source_row=row_index,
                ip_address=ip_address,
                hostname=row_data.get("hostname"),
                interface_name=row_data.get("interface_name"),
                role=row_data.get("role"),
                status=(row_data.get("status") or "assigned").lower(),
                gateway=normalize_ip(row_data.get("gateway")) or row_data.get("gateway") or (network.gateway if network else None),
                dns_servers=split_multi_value(row_data.get("dns_servers")) or (network.dns_servers if network else []),
                search_domains=split_multi_value(row_data.get("search_domains")) or (network.search_domains if network else []),
                description=row_data.get("notes") or row_data.get("broadcast_address") or row_data.get("label"),
                source_payload=row_data,
            )
            session.add(assignment)
            imported_rows += 1

        summaries.append(
            SheetImportSummary(
                sheet_name=worksheet.title,
                header_row=header_row,
                headers=header_values,
                total_rows=max(worksheet.max_row - header_row, 0),
                imported_rows=imported_rows,
                skipped_rows=skipped_rows,
            )
        )

    batch.total_rows = sum(item.total_rows for item in summaries)
    batch.imported_rows = sum(item.imported_rows for item in summaries)
    batch.notes = {
        "sheet_summaries": [
            {
                "sheet_name": item.sheet_name,
                "header_row": item.header_row,
                "total_rows": item.total_rows,
                "imported_rows": item.imported_rows,
                "skipped_rows": item.skipped_rows,
            }
            for item in summaries
        ]
    }
    return ImportSummary(batch=batch, sheet_summaries=summaries)


def import_summary_as_text(summary: ImportSummary) -> str:
    payload = {
        "filename": summary.batch.filename,
        "imported_rows": summary.imported_rows,
        "total_rows": summary.total_rows,
        "sheets": [item.__dict__ for item in summary.sheet_summaries],
    }
    return json.dumps(payload, indent=2)
